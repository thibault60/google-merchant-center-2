import streamlit as st
import requests
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font
from io import BytesIO
import re
from decimal import Decimal, InvalidOperation
from collections import defaultdict

"""
Audit d'un flux Google Merchant – version « mapping FR ➜ EN »
=============================================================
• Télécharge ou lit un XML.
• Détecte si flux Merchant (<item>) ou interne français (<Sheet1>).
• Extrait les produits, convertit les attributs FR ➜ EN.
• Normalise prix, GTIN, dimensions.
• Valide attributs clés + certification/dimensions.
• Exporte un rapport Excel, avec récap par statut (Mandatory/Recommended…).
"""

# ---------------------------------------------------------------------------
# 1)  Helpers visuels
# ---------------------------------------------------------------------------

def add_custom_css():
    st.markdown(
        """
        <style>
        body {background-color:#f8f9fa;font-family:Arial, sans-serif;}
        .main-title {color:#343a40;text-align:center;font-size:2.4rem;margin-bottom:1rem;}
        </style>
        """,
        unsafe_allow_html=True,
    )

# ---------------------------------------------------------------------------
# 2)  Téléchargement / parsing XML
# ---------------------------------------------------------------------------

def fetch_xml(url: str) -> bytes | None:
    try:
        r = requests.get(url, timeout=15)
        r.raise_for_status()
        return r.content
    except requests.exceptions.RequestException as exc:
        st.error(f"Erreur téléchargement : {exc}")
        return None

def parse_xml(content: bytes) -> ET.Element | None:
    try:
        return ET.fromstring(content)
    except ET.ParseError as exc:
        st.error(f"Erreur de parsing XML : {exc}")
        return None

# ---------------------------------------------------------------------------
# 3)  Normalisations
# ---------------------------------------------------------------------------

_PRICE_RE = re.compile(r"^(\d+[.,]?\d*)(?:\s*([A-Z]{3}))?$")
_DIMENSION_RE = re.compile(r"^\d+(?:[.,]\d+)?\s?(?:mm|cm|in|kg|g)?$", re.I)

def normalize_price(raw: str) -> str:
    if not raw or raw == "MISSING":
        return "MISSING"
    m = _PRICE_RE.match(raw.strip())
    if not m:
        return raw.strip()
    amount, currency = m.groups()
    amount = amount.replace(",", ".")
    try:
        amount = f"{Decimal(amount):.2f}".rstrip("0").rstrip(".")
    except InvalidOperation:
        return raw.strip()
    return f"{amount} {currency or 'EUR'}".strip()

def normalize_gtin(raw: str) -> str:
    if not raw or raw == "MISSING":
        return "MISSING"
    try:
        val = int(Decimal(raw))
        return f"{val:013d}"
    except (InvalidOperation, ValueError):
        return raw

# ---------------------------------------------------------------------------
# 4)  Extraction produits
# ---------------------------------------------------------------------------

_NAMESPACE = {"g": "http://base.google.com/ns/1.0"}

def analyze_products(root: ET.Element) -> list[dict]:
    merchant_items = root.findall(".//item", _NAMESPACE)
    if merchant_items:
        return [_parse_google_item(it) for it in merchant_items]
    fr_items = root.findall(".//Sheet1")
    return [_parse_french_item(it) for it in fr_items]

# ----------------------  4a) Merchant  ----------------------

def _parse_google_item(item: ET.Element) -> dict:
    g = lambda tag: (
        (item.findtext(f"g:{tag}", namespaces=_NAMESPACE) or "").strip()
        or "MISSING"
    )
    g_or_plain = lambda tag: (
        (
            item.findtext(f"g:{tag}", namespaces=_NAMESPACE)
            or item.findtext(tag)
            or ""
        ).strip()
        or "MISSING"
    )

    # Shipping (bloc)
    shipping_elem = item.find("g:shipping", _NAMESPACE)
    shipping_block = "".join(shipping_elem.itertext()).strip() if shipping_elem is not None else "MISSING"

    # product_detail peut être multiple – on concatène proprement
    product_detail_elems = item.findall("g:product_detail", _NAMESPACE)
    if product_detail_elems:
        product_detail_txt = " | ".join((" ".join(pd.itertext()).strip() for pd in product_detail_elems)).strip() or "MISSING"
    else:
        product_detail_txt = g("product_detail")

    return {
        "id": g("id"),
        "title": g_or_plain("title"),
        "description": g_or_plain("description"),
        "link": g_or_plain("link"),
        "image_link": g("image_link"),
        "price": normalize_price(g("price")),
        "sale_price": normalize_price(g("sale_price")),
        "availability": g("availability"),
        "condition": g("condition"),
        "brand": g("brand"),
        "gtin": normalize_gtin(g("gtin")),
        "mpn": g("mpn"),
        "color": g("color"),
        "size": g("size"),
        "age_group": g("age_group"),
        "gender": g("gender"),
        "item_group_id": g("item_group_id"),
        # certification & dimensions
        "certification_authority": g("certification_authority"),
        "certification_name": g("certification_name"),
        "certification_code": g("certification_code"),
        "product_length": g("product_length"),
        "product_width": g("product_width"),
        "product_height": g("product_height"),
        "product_weight": g("product_weight"),
        "shipping_length": g("shipping_length"),
        "shipping_width": g("shipping_width"),
        "shipping_height": g("shipping_height"),
        # divers
        "shipping": shipping_block,
        "shipping_weight": g("shipping_weight"),
        "pattern": g("pattern"),
        "material": g("material"),
        "additional_image_link": g("additional_image_link"),
        "size_type": g("size_type"),
        "size_system": g("size_system"),
        "canonical_link": g("canonical_link"),
        "expiration_date": g("expiration_date"),
        "sale_price_effective_date": g("sale_price_effective_date"),
        "product_highlight": g("product_highlight"),
        "ships_from_country": g("ships_from_country"),
        "minimum_handling_time": g("minimum_handling_time"),
        "max_handling_time": g("max_handling_time"),
        "availability_date": g("availability_date"),
        "product_detail": product_detail_txt,
        "google_product_category": g_or_plain("google_product_category"),
    }

# ----------------------  4b) Flux interne FR  ----------------------

FR_TO_EN_MAPPING = {
    "titre": "title",
    "identifiant": "id",
    "prix": "price",
    "prixsoldé": "sale_price",
    "prixsolde": "sale_price",
    "état": "condition",
    "etat": "condition",
    "disponibilité": "availability",
    "disponibilite": "availability",
    "lien": "link",
    "lienimage": "image_link",
    "couleur": "color",
    "taille": "size",
    "tranche_age": "age_group",
    "genre": "gender",
    "marque": "brand",
    "catégoriedeproduitsgoogle": "google_product_category",
    "categoriedeproduitsgoogle": "google_product_category",
    # dimensions / poids
    "longueur_produit": "product_length",
    "largeur_produit": "product_width",
    "hauteur_produit": "product_height",
    "poids_produit": "product_weight",
    "longueur_colis": "shipping_length",
    "largeur_colis": "shipping_width",
    "hauteur_colis": "shipping_height",
    "poids_expedition": "shipping_weight",
    # certification
    "autorité_certification": "certification_authority",
    "autorite_certification": "certification_authority",
    "nom_certification": "certification_name",
    "code_certification": "certification_code",
    # délai de traitement
    "delai_traitement_minimum": "minimum_handling_time",
    "delai_traitement_maximum": "max_handling_time",
}

def _parse_french_item(item: ET.Element) -> dict:
    data: dict = {}

    # 1) mapping direct FR -> EN
    for fr_tag, en_key in FR_TO_EN_MAPPING.items():
        raw = (item.findtext(fr_tag) or "").strip()
        if raw:
            data[en_key] = raw

    # 2) balises déjà conformes (gtin etc.)
    direct_tags = [
        "gtin", "description", "shipping", "item_group_id", "mpn",
        "pattern", "material", "additional_image_link", "size_type",
        "size_system", "canonical_link", "expiration_date",
        "sale_price_effective_date", "product_highlight",
        "ships_from_country", "minimum_handling_time", "max_handling_time",
        "availability_date", "product_detail",
        "google_product_category",
    ]
    for tag in direct_tags:
        txt = (item.findtext(tag) or "").strip()
        if txt:
            data[tag] = txt

    # 3) certification concaténée (fallback rare)
    concat_val = (item.findtext("certificationcertificationauthoritycertificationcodecertificationname") or "").strip()
    if concat_val:
        parts = concat_val.split(":")
        if len(parts) == 3:
            data["certification_authority"], data["certification_code"], data["certification_name"] = parts
        else:
            data["certification_authority"] = concat_val

    # 4) normalisations
    data["price"] = normalize_price(data.get("price", "MISSING"))
    data["sale_price"] = normalize_price(data.get("sale_price", "MISSING"))
    data["gtin"] = normalize_gtin(data.get("gtin", "MISSING"))

    # valeurs manquantes
    for key in _PRODUCT_ATTRS:
        data.setdefault(key, "MISSING")

    return data

# ---------------------------------------------------------------------------
# 5)  Validation
# ---------------------------------------------------------------------------

_PRICE_VALID_RE = re.compile(r"^\d+(?:\.\d{1,2})?\s?[A-Z]{3}$")

def validate_products(products: list[dict]) -> list[dict]:
    seen_ids: set[str] = set()
    validated: list[dict] = []

    for prod in products:
        pid = prod["id"]
        price = prod["price"]
        desc  = prod["description"]

        def missing(attr: str) -> bool:
            return prod.get(attr) in ("", "MISSING")

        dims_invalid = any(
            not _DIMENSION_RE.match(prod[attr]) for attr in (
                "product_length", "product_width", "product_height",
                "shipping_length", "shipping_width", "shipping_height",
                "product_weight",
            ) if not missing(attr)
        )

        errors = {
            "duplicate_id":                 "Erreur" if pid in seen_ids else "OK",
            "invalid_or_missing_price":     "Erreur" if price == "MISSING" or not _PRICE_VALID_RE.match(price) else "OK",
            "null_price":                   "Erreur" if price.startswith("0") else "OK",
            "missing_title":                "Erreur" if prod["title"] == "MISSING" else "OK",
            "description_missing_or_short": "Erreur" if len(desc) < 20 else "OK",
            "invalid_availability":         "Erreur" if prod["availability"] == "MISSING" else "OK",
            "missing_or_empty_color":       "Erreur" if missing("color") else "OK",
            "missing_or_empty_gender":      "Erreur" if missing("gender") else "OK",
            "missing_or_empty_size":        "Erreur" if missing("size") else "OK",
            "missing_or_empty_age_group":   "Erreur" if missing("age_group") else "OK",
            "missing_or_empty_image_link":  "Erreur" if missing("image_link") else "OK",
            # certification / dimensions
            "missing_certification":        "Erreur" if any(missing(a) for a in (
                                                "certification_authority", "certification_name", "certification_code"
                                              )) else "OK",
            "missing_dimensions_weight":    "Erreur" if any(missing(a) for a in (
                                                "product_length", "product_width", "product_height", "product_weight",
                                                "shipping_length", "shipping_width", "shipping_height"
                                              )) else "OK",
            "invalid_dimension_format":     "Erreur" if dims_invalid else "OK",
            # ✅ Nouveaux contrôles demandés
            "missing_google_product_category": "Erreur" if missing("google_product_category") else "OK",
            "missing_minimum_handling_time":   "Erreur" if missing("minimum_handling_time") else "OK",
        }

        validated.append({**prod, **errors})
        seen_ids.add(pid)

    return validated

# ---------------------------------------------------------------------------
# 6)  Export Excel – Statuts & Récap
# ---------------------------------------------------------------------------

# Liste des attributs exportés (ordre colonnes)
_PRODUCT_ATTRS = [
    "id", "title", "link", "image_link", "price", "sale_price",
    "description", "availability", "condition", "brand",
    "gtin", "mpn", "color", "size", "age_group", "gender",
    "item_group_id", "google_product_category",
    # certification & dimensions
    "certification_authority", "certification_name", "certification_code",
    "product_length", "product_width", "product_height", "product_weight",
    "shipping_length", "shipping_width", "shipping_height",
    # autres
    "shipping", "shipping_weight", "pattern", "material",
    "additional_image_link", "size_type", "size_system", "canonical_link",
    "expiration_date", "sale_price_effective_date", "product_highlight",
    "ships_from_country", "minimum_handling_time", "max_handling_time",
    "availability_date", "product_detail",
]

# Colonnes de validation
_VALIDATION_ATTRS = [
    "duplicate_id", "invalid_or_missing_price", "null_price", "missing_title",
    "description_missing_or_short", "invalid_availability", "missing_or_empty_color",
    "missing_or_empty_gender", "missing_or_empty_size", "missing_or_empty_age_group",
    "missing_or_empty_image_link", "missing_certification",
    "missing_dimensions_weight", "invalid_dimension_format",
    # ✅ nouveaux checks
    "missing_google_product_category", "missing_minimum_handling_time",
]

_HEADERS = _PRODUCT_ATTRS + _VALIDATION_ATTRS

# Table des statuts (ta liste)
FIELD_STATUS = {
    "id": "Mandatory",
    "title": "Mandatory",
    "link": "Mandatory",
    "image_link": "Mandatory",
    "price": "Mandatory",
    "description": "Mandatory",
    "availability": "Mandatory",
    "condition": "Mandatory IF SECOND-HAND or RECONDITIONED",
    "brand": "Mandatory",
    "gtin": "Mandatory if GTIN existing",
    "mpn": "Mandatory if GTIN already exists",
    "color": "Mandatory if CLOTHING AND ACCESSORIES",
    "size": "Mandatory if CLOTHING AND ACCESSORIES",
    "age_group": "Mandatory if CLOTHING AND ACCESSORIES",
    "gender": "Mandatory if CLOTHING AND ACCESSORIES",
    "item_group_id": "Mandatory FOR VARIANTS",
    "shipping": "Mandatory if not configured in GMC",
    "shipping_weight": "Recommended",
    "pattern": "Recommended if CLOTHING AND ACCESSORIES",
    "material": "Recommended if CLOTHING AND ACCESSORIES",
    "additional_image_link": "Recommended",
    "size_type": "Recommended if CLOTHING AND ACCESSORIES",
    "size_system": "Recommended if CLOTHING AND ACCESSORIES",
    "canonical_link": "Recommended",
    "expiration_date": "Recommended to stop displaying a product",
    "sale_price": "Recommended for promotions",
    "sale_price_effective_date": "Recommended for promotions",
    "product_highlight": "Recommended",
    "ships_from_country": "Recommended",
    "minimum_handling_time": "Recommended",
    "max_handling_time": "Recommended",
    "availability_date": "Mandatory if product is pre-ordered",
    "certification_authority": "Recommended for ELECTRONICS & HOUSEHOLD EQUIPEMENT",
    "certification_name": "Recommended for ELECTRONICS & HOUSEHOLD EQUIPEMENT",
    "certification_code": "Recommended for ELECTRONICS & HOUSEHOLD EQUIPEMENT",
    "product_detail": "Recommended",
    "product_length": "Recommended for ELECTRONICS & HOUSEHOLD EQUIPEMENT",
    "product_width": "Recommended for ELECTRONICS & HOUSEHOLD EQUIPEMENT",
    "product_height": "Recommended for ELECTRONICS & HOUSEHOLD EQUIPEMENT",
    "product_weight": "Recommended for ELECTRONICS & HOUSEHOLD EQUIPEMENT",
    "shipping_length": "Recommended for ELECTRONICS & HOUSEHOLD EQUIPEMENT",
    "shipping_width": "Recommended for ELECTRONICS & HOUSEHOLD EQUIPEMENT",
    "shipping_height": "Recommended for ELECTRONICS & HOUSEHOLD EQUIPEMENT",
    "google_product_category": "Mandatory",
}

def generate_excel(data: list[dict]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Validation"

    # Feuille 1 : données + flags
    ws.append(_HEADERS)
    for c in ws[1]:
        c.font = Font(bold=True)
    for prod in data:
        ws.append([prod.get(col, "") for col in _HEADERS])

    # Feuille 2 : récap par attribut
    recap = wb.create_sheet("Recap_Attributs")
    recap.append(["Attribut", "Statut", "Présents", "Manquants", "% manquant"])
    for c in recap[1]:
        c.font = Font(bold=True)

    total = len(data) or 1
    # pour synthèse par statut
    by_status_counts = defaultdict(list)  # status -> [missing_rate_of_attr1, attr2, ...]

    for attr in _PRODUCT_ATTRS:
        missing = sum(1 for p in data if p.get(attr) in ("", "MISSING"))
        status = FIELD_STATUS.get(attr, "")
        missing_pct = (missing / total) * 100
        recap.append([attr, status, total - missing, missing, f"{missing_pct:.1f}"])
        if status:
            by_status_counts[status].append(100 - missing_pct)  # on stocke la complétion pour la moyenne

    # Feuille 3 : synthèse par statut
    synth = wb.create_sheet("Synthese_par_statut")
    synth.append(["Statut", "Nb attributs", "Taux de complétion moyen (%)", "Attributs"])
    for c in synth[1]:
        c.font = Font(bold=True)

    for status, completion_list in by_status_counts.items():
        attrs = [a for a, s in FIELD_STATUS.items() if s == status and a in _PRODUCT_ATTRS]
        avg_completion = sum(completion_list) / len(completion_list) if completion_list else 0.0
        synth.append([status, len(attrs), f"{avg_completion:.1f}", ", ".join(attrs)])

    # Feuille 4 : règles (tableau brut des statuts fournis)
    rules = wb.create_sheet("Regles_Attributs")
    rules.append(["Field Name", "Status"])
    for c in rules[1]:
        c.font = Font(bold=True)
    # on réinscrit la table pour transparence
    for field, status in FIELD_STATUS.items():
        rules.append([field, status])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ---------------------------------------------------------------------------
# 7)  Interface Streamlit
# ---------------------------------------------------------------------------

def main():
    add_custom_css()
    st.markdown("<h1 class='main-title'>Audit Flux Google Merchant</h1>", unsafe_allow_html=True)

    url = st.text_input("Entrez l'URL du flux XML :")
    uploaded_file = st.file_uploader("… ou téléchargez un fichier XML :", type=["xml"])

    if st.button("Auditer le flux"):
        content = None
        if url:
            content = fetch_xml(url)
        elif uploaded_file is not None:
            content = uploaded_file.read()

        if not content:
            st.warning("Veuillez fournir une URL ou un fichier XML.")
            st.stop()

        root = parse_xml(content)
        if not root:
            st.stop()

        products = analyze_products(root)
        validated = validate_products(products)
        xlsx = generate_excel(validated)

        st.success(f"Audit terminé : {len(products)} produit(s) analysé(s).")
        st.download_button(
            "Télécharger le rapport Excel",
            data=xlsx,
            file_name="audit_flux_google_merchant.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

if __name__ == "__main__":
    main()
