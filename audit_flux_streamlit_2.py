import streamlit as st
import pandas as pd
import json
import re
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Merchant Feed Extractor", page_icon="🛒", layout="wide")

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;700&display=swap');
    .block-container { max-width: 1200px; padding-top: 2rem; }
    html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }
    .main-title { font-size: 2rem; font-weight: 700; color: #1a1a2e; margin-bottom: 0.2rem; }
    .subtitle { font-size: 1rem; color: #6b7280; margin-bottom: 2rem; }
    .metric-card {
        background: linear-gradient(135deg, #eff6ff, #dbeafe);
        border: 1px solid #bfdbfe; border-radius: 12px;
        padding: 1.2rem; text-align: center;
    }
    .metric-card h3 { font-size: 1.8rem; color: #1e40af; margin: 0; }
    .metric-card p { font-size: 0.82rem; color: #4b5563; margin: 0; }
    .stDownloadButton > button {
        background-color: #1e40af !important; color: white !important;
        border-radius: 8px !important; padding: 0.6rem 2rem !important;
        font-weight: 600 !important; width: 100% !important;
    }
    .tag { display:inline-block; padding:2px 10px; border-radius:99px; font-size:0.78rem; font-weight:600; margin:2px; }
    .tag-blue { background:#dbeafe; color:#1e40af; }
    .tag-brown { background:#fef3c7; color:#92400e; }
    .tag-green { background:#d1fae5; color:#065f46; }
</style>
""", unsafe_allow_html=True)

st.markdown('<div class="main-title">🛒 Merchant Feed Extractor</div>', unsafe_allow_html=True)
st.markdown('<div class="subtitle">Feedonomics / MaxWarehouse — Extrait les attributs du flux en colonnes Excel structurées</div>', unsafe_allow_html=True)

# ── Helpers ──
def extract_metafields(json_str):
    try:
        items = json.loads(json_str)
        return {item["namespace"] + "." + item["key"]: item["value"] for item in items if isinstance(item, dict)}
    except Exception:
        return {}

def extract_publications(json_str):
    try:
        items = json.loads(json_str)
        return " | ".join(i["name"] for i in items if isinstance(i, dict) and "name" in i)
    except Exception:
        return ""

def extract_variant_names(json_str):
    try:
        d = json.loads(json_str)
        return " | ".join(f"{k}: {v}" for k, v in d.items())
    except Exception:
        return ""

def clean_description(text):
    if not isinstance(text, str):
        return ""
    text = re.sub(r"\* ", "", text)
    text = re.sub(r"\n+", " / ", text)
    return text.strip(" /")

# ── Sidebar ──
with st.sidebar:
    st.header("⚙️ Configuration")
    separator = st.selectbox("Séparateur", ["Virgule (,)", "Tab (\\t)", "Point-virgule (;)"], index=0)
    sep_map = {"Virgule (,)": ",", "Tab (\\t)": "\t", "Point-virgule (;)": ";"}
    sep_char = sep_map[separator]

    st.divider()
    st.markdown("**Options d'extraction**")
    opt_meta     = st.checkbox("Extraire les metafields", value=True)
    opt_pub      = st.checkbox("Décomposer publications", value=True)
    opt_variants = st.checkbox("Nettoyer variant_names", value=True)
    opt_desc     = st.checkbox("Nettoyer descriptions", value=True)
    opt_es       = st.checkbox("Inclure colonnes ES (espagnol)", value=False)

    st.divider()
    st.markdown("**Légende couleurs Excel**")
    st.markdown('<span class="tag tag-blue">🔵 Attributs MC</span>', unsafe_allow_html=True)
    st.markdown('<span class="tag tag-brown">🟤 Metafields</span>', unsafe_allow_html=True)
    st.markdown('<span class="tag tag-green">🟢 Espagnol</span>', unsafe_allow_html=True)

# ── Upload ──
uploaded = st.file_uploader("📂 Importe ton flux Feedonomics (CSV / TSV)", type=["csv", "tsv", "txt"])

if uploaded:
    with st.spinner("Lecture du fichier..."):
        try:
            df = pd.read_csv(uploaded, sep=sep_char, dtype=str, encoding="utf-8-sig", on_bad_lines="skip", low_memory=False)
            df.columns = df.columns.str.strip().str.strip('"').str.lower()
            df = df.fillna("")
        except Exception as e:
            st.error(f"Erreur de lecture : {e}")
            st.stop()

    total_rows, total_cols = df.shape

    # ── Transformations ──
    with st.spinner("Transformation des colonnes..."):

        if opt_desc and "description" in df.columns:
            df["description"] = df["description"].apply(clean_description)
        if opt_desc and "es_body_html" in df.columns:
            df["es_body_html"] = df["es_body_html"].apply(clean_description)

        if opt_pub and "publications" in df.columns:
            df["publications_names"] = df["publications"].apply(extract_publications)

        if opt_variants and "variant_names" in df.columns:
            df["variant_names_clean"] = df["variant_names"].apply(extract_variant_names)

        meta_cols = []
        if opt_meta and "product_meta" in df.columns:
            meta_extracted = df["product_meta"].apply(extract_metafields)
            meta_df = pd.DataFrame(meta_extracted.tolist()).fillna("")
            meta_df.columns = ["meta." + c for c in meta_df.columns]
            meta_cols = list(meta_df.columns)
            df = pd.concat([df.reset_index(drop=True), meta_df.reset_index(drop=True)], axis=1)

    # ── Construction colonnes finales ──
    CORE_COLS = [
        "id", "item_group_id", "sku", "gtin",
        "parent_title", "child_title",
        "brand", "product_type",
        "price", "sale_price", "availability",
        "inventory_quantity", "inventory_management", "inventory_policy",
        "link", "image_link", "additional_image_link",
        "color", "size", "material",
        "weight", "weight_unit", "shipping_weight",
        "taxable", "requires_shipping", "fulfillment_service",
        "published_status", "tags",
        "custom_collections_title", "smart_collections_title",
    ]
    present_core = [c for c in CORE_COLS if c in df.columns]

    extra = []
    if "publications_names" in df.columns:  extra.append("publications_names")
    if "variant_names_clean" in df.columns: extra.append("variant_names_clean")
    if "description" in df.columns:         extra.append("description")

    es_cols = [c for c in df.columns if c.startswith("es_")] if opt_es else []

    all_cols = present_core + extra + meta_cols + es_cols
    seen = set()
    final_cols = [c for c in all_cols if not (c in seen or seen.add(c))]
    df_out = df[final_cols]

    # ── Métriques ──
    st.markdown("---")
    c1, c2, c3, c4 = st.columns(4)
    with c1: st.markdown(f'<div class="metric-card"><h3>{total_rows:,}</h3><p>Produits</p></div>', unsafe_allow_html=True)
    with c2: st.markdown(f'<div class="metric-card"><h3>{len(present_core)}</h3><p>Attributs MC</p></div>', unsafe_allow_html=True)
    with c3: st.markdown(f'<div class="metric-card"><h3>{len(meta_cols)}</h3><p>Metafields</p></div>', unsafe_allow_html=True)
    with c4: st.markdown(f'<div class="metric-card"><h3>{len(final_cols)}</h3><p>Colonnes totales</p></div>', unsafe_allow_html=True)

    st.markdown("---")

    # ── Tabs aperçu ──
    tab1, tab2, tab3 = st.tabs(["📋 Attributs principaux", "🔍 Metafields", "📊 Toutes les colonnes"])

    with tab1:
        preview_cols = [c for c in present_core + extra if c in df_out.columns]
        st.dataframe(df_out[preview_cols].head(100), use_container_width=True, height=380)

    with tab2:
        if meta_cols:
            id_col = ["id"] if "id" in df_out.columns else []
            st.dataframe(df_out[id_col + meta_cols].head(100), use_container_width=True, height=380)
        else:
            st.info("Aucun metafield extrait (option désactivée ou colonne absente).")

    with tab3:
        st.dataframe(df_out.head(50), use_container_width=True, height=380)

    # ── Build XLSX ──
    st.markdown("---")
    st.subheader("📥 Export")

    with st.spinner("Génération du fichier Excel..."):
        HDR_BLUE  = "1e3a8a"
        HDR_BROWN = "92400e"
        HDR_GREEN = "065f46"
        EVEN_BLUE  = "EFF6FF"
        EVEN_AMBER = "fffbeb"
        EVEN_GREEN = "f0fdf4"

        thin   = Side(style="thin", color="D1D5DB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def col_hdr_color(col):
            if col.startswith("meta."): return HDR_BROWN
            if col.startswith("es_"):   return HDR_GREEN
            return HDR_BLUE

        def col_row_fill(col, ri):
            even = ri % 2 == 0
            if col.startswith("meta."): return EVEN_AMBER if even else "fef9c3"
            if col.startswith("es_"):   return EVEN_GREEN if even else "f7fef9"
            return EVEN_BLUE if even else "FFFFFF"

        COL_WIDTHS = {
            "id":14,"item_group_id":14,"sku":14,"gtin":16,
            "parent_title":45,"child_title":45,"brand":16,"product_type":20,
            "price":10,"sale_price":10,"availability":14,
            "inventory_quantity":12,"inventory_management":18,"inventory_policy":16,
            "link":60,"image_link":55,"additional_image_link":40,
            "color":14,"size":12,"material":16,
            "weight":10,"weight_unit":10,"shipping_weight":14,
            "published_status":14,"tags":35,
            "custom_collections_title":35,"smart_collections_title":35,
            "publications_names":50,"variant_names_clean":25,"description":70,
        }

        MAX_ROWS = 1_048_575
        wb = Workbook()
        if "Sheet" in wb.sheetnames: del wb["Sheet"]

        sheet_idx   = 1
        batch       = []
        sheets_made = 0

        def flush(rows, name):
            ws = wb.create_sheet(title=name)
            for ci, col in enumerate(final_cols, 1):
                cell = ws.cell(row=1, column=ci, value=col)
                cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
                cell.fill      = PatternFill("solid", fgColor=col_hdr_color(col))
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border    = border
            ws.row_dimensions[1].height = 32
            ws.freeze_panes = "A2"

            data_font = Font(name="Arial", size=10)
            for ri, row in enumerate(rows, 2):
                for ci, (col, val) in enumerate(zip(final_cols, row), 1):
                    cell = ws.cell(row=ri, column=ci, value=val)
                    cell.font      = data_font
                    cell.fill      = PatternFill("solid", fgColor=col_row_fill(col, ri))
                    cell.border    = border
                    cell.alignment = Alignment(vertical="center")

            for ci, col in enumerate(final_cols, 1):
                w = COL_WIDTHS.get(col, 22 if col.startswith("meta.") else 35 if col.startswith("es_") else 18)
                ws.column_dimensions[get_column_letter(ci)].width = min(w, 80)

        for _, row_data in df_out.iterrows():
            batch.append(list(row_data))
            if len(batch) >= MAX_ROWS:
                name = f"Data_{sheet_idx}"
                flush(batch, name)
                sheet_idx += 1
                sheets_made += 1
                batch = []

        if batch:
            flush(batch, "Data" if sheet_idx == 1 else f"Data_{sheet_idx}")

        # Feuille légende
        ws_l = wb.create_sheet("Légende")
        legend = [("Attributs Merchant Center", HDR_BLUE), ("Metafields (meta.*)", HDR_BROWN), ("Espagnol (es_*)", HDR_GREEN)]
        ws_l["A1"] = "Couleur"; ws_l["B1"] = "Type"
        for c in ws_l["1:1"]: c.font = Font(name="Arial", bold=True)
        for ri, (label, clr) in enumerate(legend, 2):
            ws_l.cell(row=ri, column=1).fill = PatternFill("solid", fgColor=clr)
            ws_l.cell(row=ri, column=2, value=label).font = Font(name="Arial", size=10)
        ws_l.column_dimensions["A"].width = 5
        ws_l.column_dimensions["B"].width = 35

        output = BytesIO()
        wb.save(output)
        output.seek(0)

    fname = uploaded.name.rsplit(".", 1)[0] + "_extracted.xlsx"
    st.download_button(
        label=f"⬇️ Télécharger {fname}",
        data=output,
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

else:
    st.markdown("---")
    st.info("👆 Importe ton fichier flux Feedonomics pour commencer.")
    with st.expander("📖 Ce que fait ce script"):
        st.markdown("""
- **Attributs MC** : id, sku, gtin, titres, prix, dispo, liens images, couleur, taille, stock…
- **Metafields** : le champ `product_meta` (JSON) est décomposé en colonnes `meta.namespace.key`
- **Publications** : le JSON `publications` devient une liste lisible `Online Store | Google & YouTube…`
- **Descriptions** : nettoyage du balisage markdown
- **Colonnes ES** : traductions espagnol incluses en option
        """)
